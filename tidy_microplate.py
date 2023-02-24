# %%
# pylint: disable=invalid-name, missing-module-docstring
import argparse
import re
import sys
from datetime import datetime
from warnings import warn
from typing import TYPE_CHECKING

import openpyxl
import pandas as pd

if (sys.version_info < (3, 9)) | TYPE_CHECKING:
    warn("This script requires Python 3.9 or higher")
    sys.exit(1)


def main() -> None:
    """Main function to tidy the microplate data from the Excel file."""

    parser = argparse.ArgumentParser()
    parser.add_argument("filename", help="Filename of the Excel file to tidy")
    parser.add_argument(
        "-o",
        "--output",
        help="Output filename. Otherwise, will use the same filename as the input file",
    )
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.filename, read_only=True, data_only=True)
    all_data = tidy_microplate(wb)
    df = create_dataframe(all_data)

    if args.output:
        output = args.output
    else:
        output = args.filename.replace(".xlsx", ".csv")

    print(f"Writing to {output}")
    df.to_csv(output, index=False)


def tidy_microplate(wb: openpyxl.Workbook) -> list[dict]:
    """
    Tidy the microplate data from the Excel file.
    """
    all_data = []
    print(f"Found {len(wb.sheetnames)} sheets: {wb.sheetnames}")

    for sheet in wb:
        metadata = {}
        wells = contents = groups = day = None  # type: ignore

        for row in sheet.iter_rows(values_only=True):
            c0 = row[0]

            # Skip empty cells
            if c0 is None:
                continue

            # Add metadata
            if not re.match(r"\s?Raw Data", c0):
                # Skip miscellaneous info
                if c0.startswith(("User", "Path", "Test", "Luminescence")):
                    continue

                # Get IDs
                elif c0.startswith("ID"):
                    reg = re.search(r"^(ID\d): (.*)$", c0)
                    assert (
                        reg is not None
                    ), f"ID did not match format ID1: foo, got {c0}"
                    metadata[reg.group(1)] = reg.group(2)
                    continue

                # Get date
                elif c0.startswith("Date"):
                    reg = re.search(r"^Date: (\d{2}-\d{2}-\d{4}$)", c0)
                    assert (
                        reg is not None
                    ), f"Date did not match format dd-mm-yyyy, got {c0}"
                    day = reg.group(1)
                    continue

                # Get time
                elif c0.startswith("Time"):
                    reg = re.search(r"^Time: (\d{2}:\d{2}:\d{2}$)", c0)
                    assert (
                        reg is not None
                    ), f"Time did not match format hh:mm:ss, got {c0}"
                    metadata["date"] = datetime.strptime(
                        f"{day} {reg.group(1)}", "%d-%m-%Y %H:%M:%S"
                    )
                    continue

                # Get wells
                elif c0 == "Well":
                    wells = [c for c in row[2:]]
                    continue

                # Get contents
                elif c0 == "Content":
                    contents = [c for c in row[2:]]
                    continue

                # Get Group
                elif c0 == "Group":
                    groups = [c for c in row[2:]]
                    continue
                else:
                    raise ValueError(f"Unknown metadata {c0}")
            else:
                # Fill in data per well
                # Get channel
                if re.search(r"535 A", c0):
                    channel = "A535"
                elif re.search(r"475-30 B", c0):
                    channel = "B475"
                else:
                    raise ValueError(f"Unknown channel {c0}")

                # Get time
                t = row[1]

                # Fill in per well readouts
                assert (
                    wells is not None and contents is not None and groups is not None
                ), "Wells, contents, and groups must be defined before getting values"

                for well, sample, group, value in zip(wells, contents, groups, row[2:]):
                    data = {}

                    # Add channel and time
                    data["channel"] = channel
                    data["time"] = t

                    # Get well info
                    assert re.match(r"^[A-Z]\d+$", well), f"Invalid well {well}"
                    data["well"] = well
                    data["row"] = re.search(r"^([A-Z])", well).group(1)  # type: ignore
                    data["col"] = re.search(r"(\d+)$", well).group(1)  # type: ignore

                    # Get sample info
                    data["sample"] = sample
                    data["group"] = group

                    # Get value
                    data["value"] = value
                    all_data.append(metadata.copy() | data.copy())
    print(f"Found {len(all_data)} rows")
    return all_data


def create_dataframe(all_data: list[dict]) -> pd.DataFrame:
    """Create a tidy dataframe from the list of dictionaries
    split by channel.

    Args:
        all_data (list[dict]): list of rows as dictionaries

    Returns:
        pd.DataFrame: tidy dataframe in long format
    """
    df = pd.DataFrame(all_data)
    idx = df.columns.tolist()
    idx.remove("channel")
    idx.remove("value")
    df = df.pivot(
        index=idx,
        columns="channel",
        values="value",
    )
    df = df.reset_index(drop=False)
    df.loc[:, "BRET RATIO"] = df.loc[:, "A535"] / df.loc[:, "B475"]

    return df


if __name__ == "__main__":
    main()
